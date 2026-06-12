"""
process_netsuite_export.py
--------------------------
Parses the NetSuite "MEC - Subsidiary Due To/Froms" XLS export (XML format)
and appends clean, tagged transaction rows into the master Google Sheet
template (Due_ToFrom_Master.xlsx).

Usage:
    python process_netsuite_export.py <path_to_netsuite_export.xls>

The script:
  1. Reads the NetSuite XML export (Activity Only filter)
  2. Extracts every transaction line with full context tags
  3. Extracts per-account ending balances
  4. Opens Due_ToFrom_Master.xlsx (creates it if missing), replaces data for
     the current month, and saves.

KEY DESIGN NOTE — per-subsidiary sub balances (25001 / 15152):
  The shared sub accounts (25001 "Due to Cardone Ventures" and 15152 "Due From
  Cardone Ventures") carry balances for EVERY subsidiary.  The unfiltered
  NetSuite export only gives us an aggregate account balance.  The Recon Summary
  tab therefore:
    • Auto-computes sub-side *activity* from the Transactions tab (SUMPRODUCT)
    • Has a yellow "Sub Prior End Bal" column the user fills in once per month
    • Derives Sub End Balance = Prior + Activity
  The first time you run the script, fill in the Prior column manually from your
  existing balance sheet recon sheets.  Each subsequent month, copy the prior
  month's End Balance into the Prior column (or use IMPORTRANGE from the prior
  Google Sheet to auto-fill it).
"""

import sys
import os
import re
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XLSide
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Account pairing metadata
# ---------------------------------------------------------------------------
MIRROR_PAIRS = [
    # -------------------------------------------------------------------------
    # ADD a new subsidiary:
    #   {"label": "Short Name",  "cv": ["XXXXX"],  "sub": ["25001","15152"],
    #    "entity": "Legal Entity Name, LLC",  "kind": "shared"},
    # Use kind="dedicated" + a real sub account number when the sub has its
    # own dedicated intercompany account instead of 25001/15152.
    #
    # REMOVE a subsidiary: delete its line(s) from this list.
    # Re-run the script after any change; the Recon Summary rebuilds from scratch.
    # -------------------------------------------------------------------------
    {"label": "10X HVAC",             "cv": ["15108"],        "sub": ["25001","15152"], "entity": "10X HVAC, LLC",                                "kind": "shared"},
    {"label": "10X Buy/Sell",          "cv": ["15109","22209"],"sub": ["25001","15152"], "entity": "10X Buy/Sell, LLC",                            "kind": "shared"},
    {"label": "10X BA Franchising",    "cv": ["15113"],        "sub": ["25001","15152"], "entity": "10X Business Advisor Franchising, LLC",        "kind": "shared"},
    {"label": "5225 N Scottsdale",     "cv": ["15116"],        "sub": ["25001","15152"], "entity": "5225 N Scottsdale, LLC",                       "kind": "shared"},
    {"label": "10X Global Staffing",   "cv": ["15121"],        "sub": ["25001","15152"], "entity": "10X Global Staffing, LLC",                     "kind": "shared"},
    {"label": "10X Roofing (Due To)",  "cv": ["15129"],        "sub": ["25001","15152"], "entity": "10X Roofing Management, LLC",                  "kind": "shared"},
    {"label": "10X HomeServe",         "cv": ["15138"],        "sub": ["25001","15152"], "entity": "10X HomeServe, LLC",                           "kind": "shared"},
    {"label": "10X Coverage",          "cv": ["15143"],        "sub": ["25001","15152"], "entity": "10X Group Plan, LLC",                          "kind": "shared"},
    {"label": "CardoMax",              "cv": ["15145"],        "sub": ["25001","15152"], "entity": "CardoMax, LLC",                                "kind": "shared"},
    {"label": "10X Farms & Ranch",     "cv": ["15146"],        "sub": ["25001","15152"], "entity": "10X Farms & Ranch, LLC",                      "kind": "shared"},
    {"label": "15111 N Pima",          "cv": ["15147"],        "sub": ["25001","15152"], "entity": "15111 N Pima Road, LLC",                       "kind": "shared"},
    {"label": "HRE",                   "cv": ["15150"],        "sub": ["25001","15152"], "entity": "Heather Rae Essentials LLC",                   "kind": "shared"},
    {"label": "52CV Ventures",         "cv": ["15142"],        "sub": ["25001","15152"], "entity": "52CV Ventures, LLC",                           "kind": "shared"},
    {"label": "10X 24-7 Services",     "cv": ["15156"],        "sub": ["25001","15152"], "entity": "10X 24/7 Services, LLC",                       "kind": "shared"},
    {"label": "10X Roofing (N/R<>N/P)","cv": ["15164"],       "sub": ["23203"],         "entity": "10X Roofing Management, LLC",                  "kind": "dedicated"},
    {"label": "HRE (Prom N/R<>N/P)",   "cv": ["15204"],       "sub": ["23206"],         "entity": "Heather Rae Essentials LLC",                   "kind": "dedicated"},
    {"label": "CardoMax (Prom N/R<>N/P)","cv": ["15205"],     "sub": ["23207"],         "entity": "CardoMax, LLC",                                "kind": "dedicated"},
]

# CV account number -> counterparty entity (for tagging CV-side transactions)
CV_ACCT_ENTITY = {acct: p["entity"] for p in MIRROR_PAIRS for acct in p["cv"]}

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _clean_ns(content: str) -> str:
    content = re.sub(r' xmlns[^"]*"[^"]*"', '', content)
    content = re.sub(r'<(\w+:)', '<', content)
    content = re.sub(r'</(\w+:)', '</', content)
    content = re.sub(r' \w+:', ' ', content)
    return content


def _parse_row(row) -> list:
    cells = row.findall('.//Cell')
    vals: dict = {}
    cur = 0
    for c in cells:
        idx = c.get('Index')
        if idx:
            cur = int(idx) - 1
        data = c.find('Data')
        vals[cur] = data.text if data is not None else None
        cur += 1
    max_col = max(vals.keys()) if vals else -1
    return [vals.get(j) for j in range(max_col + 1)]


def _safe_float(val) -> float | None:
    try:
        return round(float(val), 2) if val is not None else None
    except (ValueError, TypeError):
        return None


def _parse_date(val: str | None) -> str | None:
    if not val:
        return None
    return val.split('T')[0] if 'T' in val else val


# ---------------------------------------------------------------------------
# CSV-format helpers (NetSuite now exports CSV with $-formatted amounts,
# MM/DD/YYYY dates, and a date-RANGE header e.g. "April 01, 2026 - April 30, 2026")
# ---------------------------------------------------------------------------
MONTH_ABBR = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
FULL_MONTHS = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
    'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11,
    'december': 12,
}


def _month_label(year: int, month: int) -> str:
    """(2026, 4) -> 'Apr 2026'."""
    return f"{MONTH_ABBR[month - 1]} {year}"


def _month_sort_key(label: str | None) -> tuple:
    """'Apr 2026' -> (2026, 4) for chronological sorting."""
    if not label:
        return (9999, 99)
    parts = str(label).split()
    if len(parts) == 2 and parts[0] in MONTH_ABBR:
        return (int(parts[1]), MONTH_ABBR.index(parts[0]) + 1)
    return (9999, 99)


def _parse_money(val) -> float | None:
    """'$17,053.73' -> 17053.73 ; '($17,448.73)' -> -17448.73 ; '$0.00' -> 0.0."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '-'):
        return None
    neg = s.startswith('(') and s.endswith(')')
    s = s.strip('()').replace('$', '').replace(',', '').strip()
    if s == '':
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return round(-f if neg else f, 2)


def _csv_date(val) -> tuple[str | None, str | None]:
    """'04/01/2026' -> ('2026-04-01', 'Apr 2026')."""
    if not val:
        return None, None
    s = str(val).strip()
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if not m:
        return s, None
    mo, day, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return f"{yr:04d}-{mo:02d}-{day:02d}", _month_label(yr, mo)


def _period_months(header_text: str | None) -> tuple[list[str], str | None]:
    """Parse the report header date range into a list of month labels.

    'April 01, 2026 - April 30, 2026'          -> (['Apr 2026'], 'Apr 2026')
    'March 01, 2026 - April 30, 2026'          -> (['Mar 2026','Apr 2026'], 'Apr 2026')
    Returns (all_month_labels, end_month_label).
    """
    if not header_text:
        return [], None
    dates = re.findall(r'([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})', header_text)
    parsed = []
    for name, _day, yr in dates:
        mo = FULL_MONTHS.get(name.lower())
        if mo:
            parsed.append((int(yr), mo))
    if not parsed:
        return [], None
    start, end = parsed[0], parsed[-1]
    months: list[str] = []
    y, m = start
    while (y, m) <= end:
        months.append(_month_label(y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months, _month_label(*end)


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

SKIP_STARTS = (
    'ASSETS', 'LIABILITIES', 'Current Assets', 'Current Liabilities',
    'Other Current Asset', 'Other Current Liab', 'Total',
    'Equity', 'Retained', 'Net Income', 'Options:', 'Financial Row',
)


def parse_export(filepath: str):
    """Dispatch to the CSV or XML parser based on file content/extension.

    Returns (df_trans, df_bal, meta) where:
      - df_trans / df_bal each carry a per-row 'Month' column ('Apr 2026')
      - meta = {'period_months': [...], 'bal_month': 'Apr 2026', 'format': 'csv'|'xls'}
    Transaction rows are tagged by each transaction's OWN date, so a single
    export spanning multiple months is split correctly. Account balances are
    period figures, so they are tagged to the period END month.
    """
    with open(filepath, 'rb') as f:
        head = f.read(256).lstrip()
    is_xml = head.startswith(b'<?xml') or head.startswith(b'<')

    if is_xml:
        report_month, df_t, df_b = parse_export_xml(filepath)
        return df_t, df_b, {
            'period_months': [report_month] if report_month else [],
            'bal_month': report_month,
            'format': 'xls',
        }
    return parse_export_csv(filepath)


def parse_export_csv(filepath: str):
    import csv as _csv

    with open(filepath, newline='', encoding='utf-8-sig') as fh:
        rows = list(_csv.reader(fh))

    # Locate the period header (row index 3) and the "Financial Row" column
    # header; transaction/balance data begins on the line after that header.
    period_text: str | None = None
    data_start: int | None = None
    for i, r in enumerate(rows[:15]):
        if i == 3 and r and r[0]:
            period_text = r[0].strip()
        if r and r[0] and r[0].strip().startswith('Financial Row'):
            data_start = i + 1
            break
    if data_start is None:
        raise ValueError("Could not locate the 'Financial Row' header line in CSV.")

    period_months, bal_month = _period_months(period_text)
    if not bal_month:
        raise ValueError(f"Could not parse report period from header: {period_text!r}")

    transactions: list[dict] = []
    acct_balances: list[dict] = []

    current_acct_num: str | None  = None
    current_acct_name: str | None = None
    current_acct_beg_bal: float   = 0.0

    for r in rows[data_start:]:
        def g(i):
            return r[i] if len(r) > i and r[i] not in (None, '') else None

        c0 = (r[0] or '').strip() if len(r) > 0 else ''
        c1 = (r[1] or '').strip() if len(r) > 1 else ''

        if not c0 and not c1:
            continue

        # Account header / total / section row (Financial Row filled, Type empty)
        if c0 and not c1:
            if c0.startswith('Total - '):
                if current_acct_num and current_acct_num in c0:
                    acct_balances.append({
                        'Month':       bal_month,
                        'Acct_Num':    current_acct_num,
                        'Acct_Name':   current_acct_name,
                        'Beg_Balance': current_acct_beg_bal,
                        'Activity':    _parse_money(g(5)),
                        'End_Balance': _parse_money(g(6)),
                    })
            elif any(c0.startswith(s) for s in SKIP_STARTS):
                pass
            else:
                m = re.match(r'^(\d{5})\s*-\s*(.+)$', c0)
                if m:
                    current_acct_num  = m.group(1)
                    current_acct_name = c0
                    current_acct_beg_bal = _parse_money(g(6)) or 0.0
            continue

        # Transaction row (Type filled)
        iso_date, month_label = _csv_date(g(2))
        raw_sub = g(11)
        side = "CV" if raw_sub == "Cardone Ventures, LLC" else "Sub"
        if side == "CV":
            subsidiary = CV_ACCT_ENTITY.get(
                current_acct_num,
                f"⚠ UNMAPPED CV ACCT {current_acct_num}",
            )
        else:
            subsidiary = raw_sub

        transactions.append({
            'Month':       month_label or bal_month,
            'Date':        iso_date,
            'Side':        side,
            'Subsidiary':  subsidiary,
            'Acct_Num':    current_acct_num,
            'Acct_Name':   current_acct_name,
            'Type':        g(1),
            'Doc_Number':  g(3),
            'Name':        g(4),
            'Amount':      _parse_money(g(5)),
            'Balance':     _parse_money(g(6)),
            'Description': g(7),
            'Memo':        g(8),
        })

    meta = {
        'period_months': period_months,
        'bal_month':     bal_month,
        'format':        'csv',
    }
    return pd.DataFrame(transactions), pd.DataFrame(acct_balances), meta


def parse_export_xml(filepath: str):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = _clean_ns(f.read())

    root = ET.fromstring(content)
    sheet = root.findall('.//Worksheet')[0]
    all_rows = sheet.findall('.//Row')

    # Report month is in row index 3
    report_month = None
    for i, row in enumerate(all_rows[:6]):
        r = _parse_row(row)
        if i == 3 and r and r[0]:
            report_month = r[0].strip()
    if not report_month:
        raise ValueError("Could not detect report month from file header.")

    transactions: list[dict] = []
    acct_balances: list[dict] = []

    current_acct_num: str | None  = None
    current_acct_name: str | None = None
    current_acct_beg_bal: float   = 0.0

    for row in all_rows[7:]:
        r = _parse_row(row)
        if not r or not any(v for v in r):
            continue

        first = r[0]

        # Account header / total row
        if first is not None and (len(r) < 2 or r[1] is None):
            if first.startswith('Total - '):
                end_bal  = _safe_float(r[6] if len(r) > 6 else None)
                activity = _safe_float(r[5] if len(r) > 5 else None)
                if current_acct_num:
                    acct_balances.append({
                        'Month':       report_month,
                        'Acct_Num':    current_acct_num,
                        'Acct_Name':   current_acct_name,
                        'Beg_Balance': current_acct_beg_bal,
                        'Activity':    activity,
                        'End_Balance': end_bal,
                    })
            elif any(first.startswith(s) for s in SKIP_STARTS):
                pass
            else:
                m = re.match(r'^(\d{5})\s*-\s*(.+)$', first.strip())
                if m:
                    current_acct_num  = m.group(1)
                    current_acct_name = first.strip()
                    current_acct_beg_bal = _safe_float(r[6] if len(r) > 6 else None) or 0.0
            continue

        # Transaction row
        if first is None and len(r) > 1 and r[1] is not None:
            def g(i):
                return r[i] if len(r) > i else None

            raw_sub = g(11)
            side = "CV" if raw_sub == "Cardone Ventures, LLC" else "Sub"
            # For CV-side rows NetSuite reports "Cardone Ventures, LLC" as the subsidiary.
            # Derive the actual counterparty entity from the CV account number instead.
            # Unknown CV accounts get a visible flag so they surface in the output.
            if side == "CV":
                subsidiary = CV_ACCT_ENTITY.get(
                    current_acct_num,
                    f"⚠ UNMAPPED CV ACCT {current_acct_num}",
                )
            else:
                subsidiary = raw_sub

            transactions.append({
                'Month':       report_month,
                'Date':        _parse_date(g(2)),
                'Side':        side,
                'Subsidiary':  subsidiary,
                'Acct_Num':    current_acct_num,
                'Acct_Name':   current_acct_name,
                'Type':        g(1),
                'Doc_Number':  g(3),
                'Name':        g(4),
                'Amount':      _safe_float(g(5)),
                'Balance':     _safe_float(g(6)),
                'Description': g(7),
                'Memo':        g(8),
            })

    return report_month, pd.DataFrame(transactions), pd.DataFrame(acct_balances)


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=9)
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
LIGHT_BLUE  = PatternFill("solid", fgColor="DDEEFF")

thin_side   = XLSide(border_style="thin", color="D9D9D9")
THIN_BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
MONEY_FMT   = '#,##0.00_);(#,##0.00);"-"'


def _style_cell(cell, bold=False, fill=None, align="left", number_format=None,
                font_color="000000", font_size=9):
    cell.font = Font(name="Arial", bold=bold, size=font_size, color=font_color)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format


def _header_row(ws, row_num: int, labels: list, fill=HEADER_FILL):
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row_num, column=col, value=label)
        c.font      = HEADER_FONT
        c.fill      = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN_BORDER
    ws.row_dimensions[row_num].height = 30


def _autofit(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len   = 0
        col_ltr   = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_ltr].width = min(max(max_len + 2, min_w), max_w)


# ---------------------------------------------------------------------------
# Build / update workbook
# ---------------------------------------------------------------------------

MASTER_FILE  = "Due_ToFrom_Master.xlsx"
TAB_TRANS    = "Transactions"
TAB_BAL      = "Account Balances"
TAB_RECON    = "Recon Summary"
TAB_GUIDE    = "Import Formulas"
MAX_ROWS     = 5000   # assumed upper bound for SUMPRODUCT ranges


def write_master(df_trans: pd.DataFrame,
                 df_bal:   pd.DataFrame,
                 output_path: str = MASTER_FILE):

    # Months present in the incoming data. Any existing rows for these months
    # are replaced (so re-running an export is idempotent); other months are
    # preserved. Transactions and balances can cover different month sets.
    trans_months = set(df_trans['Month'].dropna()) if not df_trans.empty else set()
    bal_months   = set(df_bal['Month'].dropna())   if not df_bal.empty   else set()

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    def get_or_create(name):
        return wb[name] if name in wb.sheetnames else wb.create_sheet(name)

    # ================================================================
    # Tab 1 — Transactions
    # ================================================================
    ws = get_or_create(TAB_TRANS)
    TRANS_HDR = ["Month","Date","Side","Subsidiary","Acct_Num","Acct_Name",
                 "Type","Doc_Number","Name","Amount","Balance","Description","Memo"]
    MONEY_T   = {10, 11}   # 1-based col indices for Amount, Balance

    existing = list(ws.iter_rows(values_only=True))
    kept = ([existing[0]] + [r for r in existing[1:] if r and r[0] not in trans_months]
            if existing and existing[0] and existing[0][0] == "Month"
            else [tuple(TRANS_HDR)])

    ws.delete_rows(1, ws.max_row + 1)
    ws.append(tuple(TRANS_HDR))
    _header_row(ws, 1, TRANS_HDR)

    for r_data in kept[1:]:
        ws.append(r_data)
        rn = ws.max_row
        for col in range(1, len(TRANS_HDR) + 1):
            c = ws.cell(rn, col)
            _style_cell(c, number_format=MONEY_FMT if col in MONEY_T else None)

    for _, row in df_trans.iterrows():
        ws.append(tuple(row[c] for c in TRANS_HDR))
        rn = ws.max_row
        for col in range(1, len(TRANS_HDR) + 1):
            c = ws.cell(rn, col)
            _style_cell(c, number_format=MONEY_FMT if col in MONEY_T else None)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit(ws)
    ws.column_dimensions['L'].width = 45
    ws.column_dimensions['M'].width = 45
    ws.column_dimensions['I'].width = 25

    # ================================================================
    # Tab 2 — Account Balances
    # ================================================================
    ws2 = get_or_create(TAB_BAL)
    BAL_HDR  = ["Month","Acct_Num","Acct_Name","Beg_Balance","Activity","End_Balance"]
    MONEY_B  = {4, 5, 6}

    existing_b = list(ws2.iter_rows(values_only=True))
    kept_b = ([existing_b[0]] + [r for r in existing_b[1:] if r and r[0] not in bal_months]
              if existing_b and existing_b[0] and existing_b[0][0] == "Month"
              else [tuple(BAL_HDR)])

    ws2.delete_rows(1, ws2.max_row + 1)
    ws2.append(tuple(BAL_HDR))
    _header_row(ws2, 1, BAL_HDR)

    for r_data in kept_b[1:]:
        ws2.append(r_data)
        rn = ws2.max_row
        for col in range(1, len(BAL_HDR) + 1):
            _style_cell(ws2.cell(rn, col), number_format=MONEY_FMT if col in MONEY_B else None)

    for _, row in df_bal.iterrows():
        ws2.append(tuple(row[c] for c in BAL_HDR))
        rn = ws2.max_row
        for col in range(1, len(BAL_HDR) + 1):
            _style_cell(ws2.cell(rn, col), number_format=MONEY_FMT if col in MONEY_B else None)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    _autofit(ws2)

    # ================================================================
    # Tab 3 — Recon Summary
    # ================================================================
    ws3 = get_or_create(TAB_RECON)

    RECON_HDR = [
        "Month",            # A  col 1
        "Relationship",     # B  col 2
        "Subsidiary",       # C  col 3
        "CV Acct(s)",       # D  col 4
        "CV End Balance",   # E  col 5  — formula from Account Balances
        "Sub Acct(s)",      # F  col 6
        "Sub Prior End Bal",# G  col 7  — YELLOW: user fills in at month-start
        "Sub Activity",     # H  col 8  — SUMPRODUCT from Transactions
        "Sub End Balance",  # I  col 9  — formula: G + H
        "Variance (CV+Sub)",# J  col 10 — formula: E + I  (should = 0)
    ]
    MONEY_R = {5, 7, 8, 9, 10}

    # Preserve prior-month data and the manual "Sub Prior End Bal" column (G)
    # We need to keep G values for ALL months (user-entered), never wipe them.
    existing_r = list(ws3.iter_rows(values_only=True))
    prior_g_vals: dict[tuple, float] = {}  # (month, relationship) -> G value

    if existing_r and existing_r[0] and existing_r[0][0] == "Month":
        hdr_r = existing_r[0]
        try:
            g_idx = list(hdr_r).index("Sub Prior End Bal")
            a_idx = 0  # Month
            b_idx = 1  # Relationship
        except ValueError:
            g_idx = None

        for r in existing_r[1:]:
            if r and r[0] and g_idx is not None:
                prior_g_vals[(r[a_idx], r[b_idx])] = r[g_idx]

    ws3.delete_rows(1, ws3.max_row + 1)
    _header_row(ws3, 1, RECON_HDR)
    ws3.row_dimensions[1].height = 32

    # Instruction row
    instruction = ("Yellow column = enter the Sub ending balance from prior month "
                   "(first month: copy from your existing recon sheets).  "
                   "Sub End Bal = Prior + Activity.  Variance should = 0 when reconciled.")
    ws3.merge_cells("A2:J2")
    instr_cell = ws3.cell(2, 1, instruction)
    instr_cell.font      = Font(name="Arial", italic=True, size=8, color="595959")
    instr_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    instr_cell.fill      = PatternFill("solid", fgColor="F2F2F2")
    ws3.row_dimensions[2].height = 30

    # Column letter helpers (1-based → letter)
    CL = {i: get_column_letter(i) for i in range(1, 11)}

    # Build the lookup keys for Account Balances tab once
    # AB!A = Month, AB!B = Acct_Num, AB!F = End_Balance
    # For multi-account CV (e.g. 15109+22209): sum SUMPRODUCT
    def cv_end_formula(row_num: int, cv_accts: list[str]) -> str:
        """SUMPRODUCT to sum end-balances across multiple CV account numbers."""
        terms = []
        for acc in cv_accts:
            terms.append(
                f"IFERROR(SUMPRODUCT(('Account Balances'!$A$2:$A${MAX_ROWS}={CL[1]}{row_num})"
                f"*('Account Balances'!$B$2:$B${MAX_ROWS}=\"{acc}\")"
                f",'Account Balances'!$F$2:$F${MAX_ROWS}),0)"
            )
        return "=" + "+".join(terms)

    def sub_activity_formula(row_num: int, sub_accts: list[str], entity: str) -> str:
        """SUMPRODUCT on Transactions: sum Amount where Month=A_row AND Subsidiary=entity AND Acct in sub_accts."""
        acct_clause = "+".join(
            f"(Transactions!$E$2:$E${MAX_ROWS}=\"{a}\")" for a in sub_accts
        )
        return (
            f"=SUMPRODUCT("
            f"(Transactions!$A$2:$A${MAX_ROWS}={CL[1]}{row_num})"
            f"*(Transactions!$D$2:$D${MAX_ROWS}=\"{entity}\")"
            f"*(({acct_clause})>0)"
            f",Transactions!$J$2:$J${MAX_ROWS})"
        )

    def sub_dedicated_end_formula(row_num: int, sub_accts: list[str]) -> str:
        """For dedicated sub accounts, pull from Account Balances directly."""
        terms = []
        for acc in sub_accts:
            terms.append(
                f"IFERROR(SUMPRODUCT(('Account Balances'!$A$2:$A${MAX_ROWS}={CL[1]}{row_num})"
                f"*('Account Balances'!$B$2:$B${MAX_ROWS}=\"{acc}\")"
                f",'Account Balances'!$F$2:$F${MAX_ROWS}),0)"
            )
        return "=" + "+".join(terms)

    # Build one 15-pair block per month we have balance data for, sorted
    # chronologically. Each formula self-references its own row's Month cell,
    # so any number of month blocks works without hardcoding row positions.
    all_months = sorted(
        {row[0] for row in ws2.iter_rows(min_row=2, values_only=True)
         if row and row[0]},
        key=_month_sort_key,
    )

    GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
    data_start = 3  # row 1 = header, row 2 = instruction
    rn = data_start
    for month in all_months:
        for pair in MIRROR_PAIRS:
            cv_display  = " + ".join(pair["cv"])
            sub_display = " + ".join(pair["sub"])
            prior_g = prior_g_vals.get((month, pair["label"]))

            _style_cell(ws3.cell(rn, 1, month))
            _style_cell(ws3.cell(rn, 2, pair["label"]), bold=True)
            _style_cell(ws3.cell(rn, 3, pair["entity"]))
            _style_cell(ws3.cell(rn, 4, cv_display), align="center")
            _style_cell(ws3.cell(rn, 5, cv_end_formula(rn, pair["cv"])),
                        number_format=MONEY_FMT, fill=LIGHT_BLUE)
            _style_cell(ws3.cell(rn, 6, sub_display), align="center")

            if pair["kind"] == "shared":
                # Col G: Sub Prior End Bal (manual yellow input)
                _style_cell(ws3.cell(rn, 7, prior_g),
                            number_format=MONEY_FMT, fill=YELLOW_FILL)
                # Col H: Sub Activity (SUMPRODUCT from Transactions)
                _style_cell(ws3.cell(rn, 8, sub_activity_formula(rn, pair["sub"], pair["entity"])),
                            number_format=MONEY_FMT, fill=LIGHT_BLUE)
                # Col I: Sub End Balance = G + H
                _style_cell(ws3.cell(rn, 9, f"={CL[7]}{rn}+{CL[8]}{rn}"),
                            number_format=MONEY_FMT, fill=LIGHT_BLUE)
            else:
                # Dedicated sub account — End Balance comes straight from Account Balances
                c = ws3.cell(rn, 7, "N/A"); _style_cell(c, align="center"); c.fill = GRAY_FILL
                c = ws3.cell(rn, 8, "N/A"); _style_cell(c, align="center"); c.fill = GRAY_FILL
                _style_cell(ws3.cell(rn, 9, sub_dedicated_end_formula(rn, pair["sub"])),
                            number_format=MONEY_FMT, fill=LIGHT_BLUE)

            # Col J: Variance = E + I (should be 0 when reconciled; CV is +, Sub is -)
            _style_cell(ws3.cell(rn, 10, f"={CL[5]}{rn}+{CL[9]}{rn}"),
                        number_format=MONEY_FMT)
            rn += 1

    ws3.freeze_panes = "A3"
    _autofit(ws3)
    ws3.column_dimensions['B'].width = 28
    ws3.column_dimensions['C'].width = 38
    ws3.column_dimensions['J'].width = 20

    # ================================================================
    # Tab 4 — Import Formulas (write once)
    # ================================================================
    if TAB_GUIDE not in wb.sheetnames:
        ws4 = wb.create_sheet(TAB_GUIDE)
        _write_guide(ws4)

    # Re-order tabs
    desired = [TAB_TRANS, TAB_BAL, TAB_RECON, TAB_GUIDE]
    for target_i, name in enumerate(desired):
        if name in wb.sheetnames:
            cur_i = wb.sheetnames.index(name)
            wb.move_sheet(name, offset=target_i - cur_i)

    wb.save(output_path)

    tx_m  = ", ".join(sorted(trans_months, key=_month_sort_key)) or "(none)"
    bal_m = ", ".join(sorted(bal_months, key=_month_sort_key)) or "(none)"
    print(f"\nSaved: {output_path}")
    print(f"  Transactions:     {len(df_trans)} rows  ({tx_m})")
    print(f"  Account Balances: {len(df_bal)} rows  ({bal_m})")
    print(f"  Recon Summary:    {len(all_months)} month(s) x {len(MIRROR_PAIRS)} pairs")
    print()
    print("NEXT STEP: Open Due_ToFrom_Master.xlsx, go to 'Recon Summary',")
    print("  and fill in the yellow 'Sub Prior End Bal' column with each")
    print("  subsidiary's prior-month ending balance for 25001/15152.")
    print("  (Copy these from your existing balance sheet recon Google Sheets.)")


# ---------------------------------------------------------------------------
# Import Formulas tab
# ---------------------------------------------------------------------------

def _write_guide(ws):
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 82
    ws.column_dimensions['C'].width = 50

    def hdr(row, text, col_span="A:C", fill=HEADER_FILL, height=26):
        ws.merge_cells(f"{col_span}{row}:{col_span.split(':')[1]}{row}")
        c = ws.cell(row, 1, text)
        c.font      = HEADER_FONT
        c.fill      = fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = height

    def row3(r, col1, col2, col3="", bold1=False, h=18):
        for col_i, val, bold in [(1, col1, bold1), (2, col2, False), (3, col3, False)]:
            c = ws.cell(r, col_i, val)
            c.font      = Font(name="Arial", bold=bold, size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = THIN_BORDER
        ws.row_dimensions[r].height = h

    r = 1
    hdr(r, "MASTER DUE TO/FROM — GOOGLE SHEETS IMPORT GUIDE", "A:C")
    r += 2

    hdr(r, "MONTHLY WORKFLOW", "A:C", fill=SUBHDR_FILL)
    r += 1
    row3(r, "Step 1 – Run script", 'python process_netsuite_export.py "MEC-SubsidiaryDueTo_Froms-XXX.xls"\nThis updates all tabs in Due_ToFrom_Master.xlsx with the new month.', "", bold1=True, h=42)
    r += 1
    row3(r, "Step 2 – Import to Google Sheets", "In Google Sheets: File > Import > Upload > Replace spreadsheet.\nOr use Drive if auto-sync is set up.", "", bold1=True, h=35)
    r += 1
    row3(r, "Step 3 – Fill yellow column", "In the Recon Summary tab, fill the yellow 'Sub Prior End Bal' column\nwith each sub's prior-month 25001/15152 ending balance.", "", bold1=True, h=35)
    r += 1
    row3(r, "Step 4 – Subsidiary sheets auto-update", "Any subsidiary sheet using IMPORTRANGE will update automatically\nas soon as the master sheet is updated.", "", bold1=True, h=35)
    r += 2

    hdr(r, "ONE-TIME: Authorize IMPORTRANGE (do this in each subsidiary sheet)", "A:C", fill=SUBHDR_FILL)
    r += 1
    row3(r, "Paste in any cell:", '=IMPORTRANGE("YOUR_MASTER_URL","Transactions!A1")\nClick "Allow access" when prompted.  Replace YOUR_MASTER_URL with the Google Sheets URL.', "", bold1=True, h=45)
    r += 2

    hdr(r, "QUERY(IMPORTRANGE()) FORMULA TEMPLATES", "A:C", fill=SUBHDR_FILL)
    r += 1
    for col_i, val in [(1,"Purpose"),(2,"Formula — replace URL and subsidiary name"),(3,"Notes")]:
        c = ws.cell(r, col_i, val)
        c.font   = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill   = PatternFill("solid", fgColor="404040")
        c.border = THIN_BORDER
    r += 1

    formulas = [
        ("CV-side transactions for one sub (month filter)",
         '=QUERY(IMPORTRANGE("URL","Transactions!A:M"),\n'
         ' "SELECT Col1,Col2,Col3,Col4,Col5,Col6,Col7,Col8,Col9,Col10,Col11,Col12,Col13\n'
         '  WHERE Col3=\'CV\' AND Col4=\'10X Roofing Management, LLC\' AND Col1=\'"&B2&"\'\n")',
         "B2 = month cell, e.g. \"Mar 2026\"", 72),
        ("Sub-side transactions for one sub (month filter)",
         '=QUERY(IMPORTRANGE("URL","Transactions!A:M"),\n'
         ' "SELECT Col1,Col2,Col3,Col4,Col5,Col6,Col7,Col8,Col9,Col10,Col11,Col12,Col13\n'
         '  WHERE Col3=\'Sub\' AND Col4=\'10X Roofing Management, LLC\' AND Col1=\'"&B2&"\'\n")',
         "Pulls from sub's books (25001 / 15152)", 72),
        ("CV account ending balance (Account Balances tab)",
         '=QUERY(IMPORTRANGE("URL","Account Balances!A:F"),\n'
         ' "SELECT Col6 WHERE Col1=\'"&B2&"\' AND Col2=\'15129\'")',
         "Replace 15129 with the relevant CV account number", 45),
        ("Full Recon Summary for one month",
         '=QUERY(IMPORTRANGE("URL","Recon Summary!A:J"),\n'
         ' "SELECT Col1,Col2,Col3,Col4,Col5,Col6,Col7,Col8,Col9,Col10\n'
         '  WHERE Col1=\'"&B2&"\'")',
         "Pulls the full reconciliation table for the month", 55),
        ("Variance check for one subsidiary",
         '=QUERY(IMPORTRANGE("URL","Recon Summary!A:J"),\n'
         ' "SELECT Col2,Col5,Col9,Col10\n'
         '  WHERE Col1=\'"&B2&"\' AND Col3=\'10X Roofing Management, LLC\'")',
         "Returns: Relationship, CV End, Sub End, Variance", 55),
        ("All sub-side transactions for one account number",
         '=QUERY(IMPORTRANGE("URL","Transactions!A:M"),\n'
         ' "SELECT Col1,Col2,Col7,Col8,Col9,Col10,Col11,Col12,Col13\n'
         '  WHERE Col3=\'Sub\' AND Col5=\'25001\' AND Col4=\'10X Roofing Management, LLC\'\n'
         '  AND Col1=\'"&B2&"\' ORDER BY Col2")',
         "Filter by Acct_Num (Col5) and Subsidiary (Col4)", 72),
    ]

    for purpose, formula, note, h in formulas:
        row3(r, purpose, formula, note, bold1=False, h=h)
        r += 1

    r += 1
    hdr(r, "TRANSACTIONS TAB — COLUMN REFERENCE", "A:C", fill=SUBHDR_FILL)
    r += 1
    col_ref = [
        ("Col1 = Month",       "Period string — e.g. \"Mar 2026\"  (filter with: WHERE Col1='Mar 2026')"),
        ("Col2 = Date",        "Transaction date as YYYY-MM-DD string"),
        ("Col3 = Side",        "\"CV\" = Cardone Ventures books  |  \"Sub\" = subsidiary's own books"),
        ("Col4 = Subsidiary",  "Legal entity name — e.g. \"10X Roofing Management, LLC\""),
        ("Col5 = Acct_Num",    "5-digit account number — e.g. \"15129\", \"25001\", \"15152\""),
        ("Col6 = Acct_Name",   "Full account name from NetSuite"),
        ("Col7 = Type",        "Transaction type: Journal, Bill, Payment, Credit Card, etc."),
        ("Col8 = Doc_Number",  "NetSuite document / journal entry number"),
        ("Col9 = Name",        "Vendor / customer / employee name"),
        ("Col10 = Amount",     "Transaction amount — positive = debit, negative = credit"),
        ("Col11 = Balance",    "Cumulative running balance for the account (whole account, not per sub)"),
        ("Col12 = Description","Transaction description / narration"),
        ("Col13 = Memo",       "Memo field"),
    ]
    for col_name, col_desc in col_ref:
        row3(r, col_name, col_desc, "", bold1=True, h=18)
        r += 1

    r += 1
    hdr(r, "SUB PRIOR END BALANCE — HOW TO FILL IT", "A:C", fill=SUBHDR_FILL)
    r += 1
    explanation = (
        "The 'Sub Prior End Bal' column (yellow, Recon Summary) is a manual input.  "
        "It represents each subsidiary's share of the 25001/15152 account balance at the "
        "START of the current month (= end of prior month).\n\n"
        "First time setup: copy these figures from the reconciliation totals in each "
        "subsidiary's existing balance sheet recon Google Sheet (the 'Reconciliation Total' "
        "cell on each 25001 / 15152 tab).\n\n"
        "Each subsequent month: the Sub End Balance from the prior month becomes the new "
        "Sub Prior End Bal.  You can automate this by using IMPORTRANGE from the prior "
        "month's Google Sheet file, or by copying the values over manually."
    )
    ws.merge_cells(f"A{r}:C{r}")
    c = ws.cell(r, 1, explanation)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border    = THIN_BORDER
    ws.row_dimensions[r].height = 100


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python process_netsuite_export.py <export.csv | export.xls> [more files ...]")
        print()
        print("Accepts the NetSuite 'MEC - Subsidiary Due To/Froms' export in either")
        print("the new CSV format or the legacy XML .xls format. Pass one file per")
        print("month for accurate per-month balances; multiple files may be given to")
        print("backfill several months in one run.")
        sys.exit(1)

    input_files = sys.argv[1:]
    script_dir  = os.path.dirname(os.path.abspath(sys.argv[0]))
    output_file = os.path.join(script_dir, MASTER_FILE)

    for input_file in input_files:
        if not os.path.exists(input_file):
            print(f"File not found: {input_file}")
            sys.exit(1)

        print(f"\nParsing: {input_file}")
        df_t, df_b, meta = parse_export(input_file)
        period = ", ".join(meta["period_months"]) or "(unknown)"
        print(f"  Format:           {meta['format']}")
        print(f"  Report period:    {period}")
        print(f"  Transactions:     {len(df_t)}  (tagged by transaction date)")
        print(f"  Account balances: {len(df_b)}  (period end: {meta['bal_month']})")

        if len(meta["period_months"]) > 1:
            print()
            print("  ** WARNING: this export spans more than one month. Transactions are")
            print("     split correctly by date, but Account Balances are period figures")
            print(f"     captured only for the END month ({meta['bal_month']}). For accurate")
            print("     per-month CV ending balances, export ONE month at a time.")

        # Warn about any CV account numbers not found in MIRROR_PAIRS
        if not df_t.empty:
            flagged = df_t[df_t['Subsidiary'].str.startswith('⚠', na=False)]
            if not flagged.empty:
                unknown = sorted(flagged['Acct_Num'].dropna().unique())
                print()
                print("  !! UNMAPPED CV ACCOUNT(S) — add these to MIRROR_PAIRS in the script:")
                for acct in unknown:
                    sample = flagged[flagged['Acct_Num'] == acct]['Acct_Name'].iloc[0]
                    print(f"     Acct {acct}  ({sample})")
                print()
                print('     Template (add inside MIRROR_PAIRS):')
                print('     {"label": "Short Name",  "cv": ["XXXXX"],')
                print('      "sub": ["25001","15152"],')
                print('      "entity": "Legal Entity Name, LLC",  "kind": "shared"},')
                print()
                print("     These rows appear as '⚠ UNMAPPED CV ACCT XXXXX' in the")
                print("     Transactions tab so you can find them easily.")

        write_master(df_t, df_b, output_file)
