from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

URL = "https://docs.google.com/spreadsheets/d/1KSKAav0X-dEs1HIh9sxBNC7-JDma0lH_QyZPT_CyePg"

PAIRS = [
    {"label": "10X HVAC",                  "cv": ["15108"],         "sub": ["25001","15152"], "entity": "10X HVAC, LLC",                                    "kind": "shared"},
    {"label": "10X Buy/Sell",              "cv": ["15109","22209"], "sub": ["25001","15152"], "entity": "10X Buy/Sell, LLC",                                "kind": "shared"},
    {"label": "10X BA Franchising",        "cv": ["15113"],         "sub": ["25001","15152"], "entity": "10X Business Advisor Franchising, LLC",            "kind": "shared"},
    {"label": "5225 N Scottsdale",         "cv": ["15116"],         "sub": ["25001","15152"], "entity": "5225 N Scottsdale, LLC",                           "kind": "shared"},
    {"label": "10X Global Staffing",       "cv": ["15121"],         "sub": ["25001","15152"], "entity": "10X Global Staffing, LLC",                         "kind": "shared"},
    {"label": "10X Roofing (Due To)",      "cv": ["15129"],         "sub": ["25001","15152"], "entity": "10X Roofing Management, LLC",                      "kind": "shared"},
    {"label": "10X HomeServe",             "cv": ["15138"],         "sub": ["25001","15152"], "entity": "10X HomeServe, LLC",                               "kind": "shared"},
    {"label": "10X Coverage",              "cv": ["15143"],         "sub": ["25001","15152"], "entity": "10X Group Plan, LLC",        "alt_entity": "10X Coverage, LLC",        "kind": "shared"},
    {"label": "CardoMax",                  "cv": ["15145"],         "sub": ["25001","15152"], "entity": "CardoMax, LLC",                                    "kind": "shared"},
    {"label": "10X Farms & Ranch",         "cv": ["15146"],         "sub": ["25001","15152"], "entity": "10X Farms & Ranch, LLC",                           "kind": "shared"},
    {"label": "15111 N Pima",              "cv": ["15147"],         "sub": ["25001","15152"], "entity": "15111 N Pima Road, LLC",                           "kind": "shared"},
    {"label": "HRE",                       "cv": ["15150"],         "sub": ["25001","15152"], "entity": "Heather Rae Essentials LLC",                       "kind": "shared"},
    {"label": "52CV Ventures",             "cv": ["15142"],         "sub": ["25001","15152"], "entity": "52CV Ventures, LLC",                               "kind": "shared"},
    {"label": "10X 24-7 Services",         "cv": ["15156"],         "sub": ["25001","15152"], "entity": "10X 24/7 Services, LLC",                           "kind": "shared"},
    {"label": "10X Roofing (N/R<>N/P)",   "cv": ["15164"],         "sub": ["23203"],         "entity": "10X Roofing Management, LLC",                      "kind": "dedicated"},
    {"label": "HRE (Prom N/R<>N/P)",      "cv": ["15204"],         "sub": ["23206"],         "entity": "Heather Rae Essentials LLC",                       "kind": "dedicated"},
    {"label": "CardoMax (Prom N/R<>N/P)", "cv": ["15205"],         "sub": ["23207"],         "entity": "CardoMax, LLC",                                    "kind": "dedicated"},
]

# Group pairs by entity, preserving order
entity_pairs = defaultdict(list)
entity_order = []
for p in PAIRS:
    if p["entity"] not in entity_order:
        entity_order.append(p["entity"])
    entity_pairs[p["entity"]].append(p)

# Sheet name map (no slashes allowed)
SHEET_NAMES = {
    "10X HVAC, LLC":                                    "10X HVAC",
    "10X Buy/Sell, LLC":                                "10X Buy-Sell",
    "10X Business Advisor Franchising, LLC":            "10X BA Franchising",
    "5225 N Scottsdale, LLC":                           "5225 N Scottsdale",
    "10X Global Staffing, LLC":                         "10X Global Staffing",
    "10X Roofing Management, LLC":                      "10X Roofing",
    "10X HomeServe, LLC":                               "10X HomeServe",
    "10X Group Plan, LLC":                              "10X Coverage",
    "CardoMax, LLC":                                    "CardoMax",
    "10X Farms & Ranch, LLC":                           "10X Farms & Ranch",
    "15111 N Pima Road, LLC":                           "15111 N Pima",
    "Heather Rae Essentials LLC":                       "HRE",
    "52CV Ventures, LLC":                               "52CV Ventures",
    "10X 24/7 Services, LLC":                           "10X 24-7 Services",
}

# ── Style helpers ──────────────────────────────────────────────────────────
BG_HEADER  = PatternFill("solid", fgColor="0F1117")  # dark
BG_SECTION = PatternFill("solid", fgColor="1C2333")  # mid-dark
BG_ROW     = PatternFill("solid", fgColor="161B24")  # row
BG_FORMULA = PatternFill("solid", fgColor="0F1117")  # formula cell

FG_WHITE  = Font(name="Arial", bold=True,  color="F1F5F9", size=11)
FG_TITLE  = Font(name="Arial", bold=True,  color="3B82F6", size=13)
FG_LABEL  = Font(name="Arial", bold=False, color="94A3B8", size=9)
FG_SECT   = Font(name="Arial", bold=True,  color="F59E0B", size=9)
FG_FORMULA= Font(name="Courier New", bold=False, color="10B981", size=9)
FG_NOTES  = Font(name="Arial", bold=False, color="475569", size=9)
FG_NOTE_H = Font(name="Arial", bold=True,  color="F1F5F9", size=9)

WRAP = Alignment(wrap_text=True, vertical="top")
VCENTER = Alignment(vertical="center")

thin = Side(style="thin", color="1C2333")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fill:      c.fill = fill
    if alignment: c.alignment = alignment
    if border:    c.border = border
    return c

def entity_filter(entity, alt=None):
    """Build the Col4 WHERE clause, adding OR for alt_entity when present."""
    if alt:
        return f"(Col4='{entity}' OR Col4='{alt}')"
    return f"Col4='{entity}'"

def build_formulas(pair):
    entity = pair["entity"]
    alt    = pair.get("alt_entity")
    label  = pair["label"]
    cv_accts = pair["cv"]
    sub_accts = pair["sub"]
    kind = pair["kind"]
    ef   = entity_filter(entity, alt)  # entity filter clause

    rows = []

    # ── Transactions ──────────────────────────────────────────────────────
    cols = "Col1,Col2,Col3,Col4,Col5,Col6,Col7,Col8,Col9,Col10,Col11,Col12,Col13"
    base = f'IMPORTRANGE("{URL}","Transactions!A:M")'

    cv_tx = (f'=QUERY({base},"SELECT {cols} '
             f"WHERE Col3='CV' AND {ef} AND Col1='\"&B3&\"'\")")
    rows.append(("CV-side transactions — current month",
                 cv_tx,
                 "B3 = month cell, e.g. \"Mar 2026\". Returns all CV-book entries for this entity in the selected month."))

    sub_tx = (f'=QUERY({base},"SELECT {cols} '
              f"WHERE Col3='Sub' AND {ef} AND Col1='\"&B3&\"'\")")
    rows.append(("Sub-side transactions — current month",
                 sub_tx,
                 "Returns 25001 / 15152 (or dedicated N/P acct) rows for this entity in the selected month."))

    all_tx = (f'=QUERY({base},"SELECT {cols} '
              f"WHERE {ef} AND Col1='\"&B3&\"'\")")
    rows.append(("All transactions (CV + Sub) — current month",
                 all_tx,
                 "Both sides combined for the selected month."))

    # Historical (up to and including B3 month)
    # Date column is stored as YYYY-MM-DD text — use plain string comparison (no date keyword)
    date_filter = '"&TEXT(EOMONTH(DATEVALUE("1 "&B3),0),"yyyy-mm-dd")&"'
    cv_hist = (f'=QUERY({base},"SELECT {cols} '
               f"WHERE Col3='CV' AND {ef} AND Col2 <= '{date_filter}'\")")
    rows.append(("CV-side transactions — all history through B3",
                 cv_hist,
                 "Returns every CV-side entry for this entity up to and including the month in B3. YYYY-MM-DD string comparison — no 'date' keyword needed."))

    sub_hist = (f'=QUERY({base},"SELECT {cols} '
                f"WHERE Col3='Sub' AND {ef} AND Col2 <= '{date_filter}'\")")
    rows.append(("Sub-side transactions — all history through B3",
                 sub_hist,
                 "Returns every Sub-side entry for this entity up to and including the month in B3."))

    # ── Account Balances ──────────────────────────────────────────────────
    bal_base = f'IMPORTRANGE("{URL}","Account Balances!A:F")'
    rows.append(("", "", ""))  # spacer

    for acct in cv_accts:
        rows.append((f"CV acct {acct} — end balance for month",
                     f'=QUERY({bal_base},"SELECT Col6 WHERE Col1=\'\"&B3&\"\'  AND Col2=\'{acct}\'")',
                     f"Returns the ending balance for account {acct} in the selected month. Single value."))

    if kind == "dedicated":
        for acct in sub_accts:
            rows.append((f"Sub acct {acct} — end balance for month",
                         f'=QUERY({bal_base},"SELECT Col6 WHERE Col1=\'\"&B3&\"\'  AND Col2=\'{acct}\'")',
                         f"Dedicated N/P intercompany account. Returns ending balance for selected month."))

    # ── Recon Summary ─────────────────────────────────────────────────────
    recon_base = f'IMPORTRANGE("{URL}","Recon Summary!A:J")'
    rows.append(("", "", ""))  # spacer

    rows.append((f"Recon row — {label} — current month",
                 f'=QUERY({recon_base},"SELECT Col1,Col2,Col5,Col7,Col8,Col9,Col10 WHERE Col1=\'\"&B3&\"\' AND Col2=\'{label}\'")',
                 "Returns: Month | Relationship | CV End Bal | Sub Prior | Sub Activity | Sub End Bal | Variance"))

    rows.append((f"Recon history — {label} — all months",
                 f'=QUERY({recon_base},"SELECT Col1,Col2,Col5,Col7,Col8,Col9,Col10 WHERE Col2=\'{label}\' ORDER BY Col1")',
                 "No month filter. Full history sorted chronologically."))

    # For entities that also appear in the Recon Summary under another label,
    # those are handled by their own pair section — no extra rows needed here.

    return rows


wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ── One sheet per subsidiary ───────────────────────────────────────────────
for entity in entity_order:
    pairs = entity_pairs[entity]
    ws = wb.create_sheet(title=SHEET_NAMES[entity])

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 120
    ws.column_dimensions["C"].width = 60
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # Title row
    set_cell(ws, 1, 1, entity, font=FG_TITLE, fill=BG_HEADER, alignment=Alignment(vertical="center"))
    ws.merge_cells("A1:C1")
    ws.cell(1,2).fill = BG_HEADER
    ws.cell(1,3).fill = BG_HEADER

    # Column headers
    for col, hdr in enumerate(["Formula Purpose", "Formula (paste into Google Sheets)", "Notes"], 1):
        set_cell(ws, 2, col, hdr, font=FG_NOTE_H, fill=BG_SECTION,
                 alignment=Alignment(horizontal="center", vertical="center"), border=BORDER)

    current_row = 3

    for pi, pair in enumerate(pairs):
        # Section header for the pair
        label_text = pair["label"]
        if pair["kind"] == "shared":
            acct_str = " + ".join(pair["cv"]) + "  <->  25001 / 15152"
        else:
            acct_str = " + ".join(pair["cv"]) + "  <->  " + " / ".join(pair["sub"])

        ws.row_dimensions[current_row].height = 18
        set_cell(ws, current_row, 1, f"{label_text}  ({acct_str})",
                 font=FG_SECT, fill=BG_SECTION,
                 alignment=Alignment(vertical="center"))
        ws.merge_cells(f"A{current_row}:C{current_row}")
        ws.cell(current_row, 2).fill = BG_SECTION
        ws.cell(current_row, 3).fill = BG_SECTION
        current_row += 1

        formula_rows = build_formulas(pair)
        for purpose, formula, notes in formula_rows:
            if not purpose and not formula:
                current_row += 1
                continue
            ws.row_dimensions[current_row].height = 42
            set_cell(ws, current_row, 1, purpose,
                     font=FG_LABEL, fill=BG_ROW,
                     alignment=Alignment(vertical="top", wrap_text=True), border=BORDER)
            set_cell(ws, current_row, 2, formula,
                     font=FG_FORMULA, fill=BG_FORMULA,
                     alignment=Alignment(vertical="top", wrap_text=True, horizontal="left"), border=BORDER)
            set_cell(ws, current_row, 3, notes,
                     font=FG_NOTES, fill=BG_ROW,
                     alignment=Alignment(vertical="top", wrap_text=True), border=BORDER)
            current_row += 1

        current_row += 1  # gap between pairs

    ws.freeze_panes = "A3"

# ── Index sheet ───────────────────────────────────────────────────────────
idx = wb.create_sheet(title="Index", index=0)
idx.column_dimensions["A"].width = 32
idx.column_dimensions["B"].width = 50
idx.column_dimensions["C"].width = 20
idx.row_dimensions[1].height = 30

set_cell(idx, 1, 1, "Due To/From Master — IMPORTRANGE Query Reference",
         font=FG_TITLE, fill=BG_HEADER, alignment=Alignment(vertical="center"))
idx.merge_cells("A1:C1")
idx.cell(1,2).fill = BG_HEADER
idx.cell(1,3).fill = BG_HEADER

for col, hdr in enumerate(["Subsidiary", "Entity Name", "Sheet"], 1):
    set_cell(idx, 2, col, hdr, font=FG_NOTE_H, fill=BG_SECTION,
             alignment=Alignment(horizontal="center", vertical="center"), border=BORDER)

idx.row_dimensions[2].height = 18

for i, entity in enumerate(entity_order, start=3):
    idx.row_dimensions[i].height = 18
    sheet_name = SHEET_NAMES[entity]
    pairs = entity_pairs[entity]
    cv_list = []
    for p in pairs:
        cv_list += p["cv"]
    set_cell(idx, i, 1, sheet_name,     font=FG_LABEL, fill=BG_ROW, border=BORDER,
             alignment=Alignment(vertical="center"))
    set_cell(idx, i, 2, entity,         font=FG_LABEL, fill=BG_ROW, border=BORDER,
             alignment=Alignment(vertical="center"))
    set_cell(idx, i, 3, ", ".join(cv_list), font=FG_LABEL, fill=BG_ROW, border=BORDER,
             alignment=Alignment(vertical="center"))

# Notes section
nr = len(entity_order) + 4
set_cell(idx, nr, 1, "HOW TO USE", font=FG_SECT, fill=BG_SECTION)
idx.merge_cells(f"A{nr}:C{nr}")
idx.cell(nr,2).fill = BG_SECTION
idx.cell(nr,3).fill = BG_SECTION

notes = [
    ("B3 reference", "All month-filtered formulas reference cell B3 as the month string (e.g. \"Mar 2026\"). Put your month label in B3 on each subsidiary sheet."),
    ("One-time auth", 'Paste =IMPORTRANGE("' + URL + '","Transactions!A1") into any cell on a new subsidiary sheet and click "Allow access" when prompted. Do this once per sheet.'),
    ("Paste as formula", "Copy a formula cell from the green column, paste into Google Sheets starting with =. The formula is plain text here — it becomes live when pasted."),
    ("Dedicated N/P pairs", "10X Roofing, HRE, and CardoMax each have TWO pairs: a Due To/From pair (25001/15152) and a Promissory Note pair (23203/23206/23207). Separate tabs cover each."),
    ("Master URL", URL),
]
for j, (k, v) in enumerate(notes, start=nr+1):
    idx.row_dimensions[j].height = 36
    set_cell(idx, j, 1, k, font=FG_NOTE_H, fill=BG_ROW, border=BORDER,
             alignment=Alignment(vertical="top", wrap_text=True))
    c = idx.cell(row=j, column=2, value=v)
    c.font = FG_NOTES
    c.fill = BG_ROW
    c.border = BORDER
    c.alignment = Alignment(vertical="top", wrap_text=True)
    idx.merge_cells(f"B{j}:C{j}")
    idx.cell(j,3).fill = BG_ROW

idx.freeze_panes = "A3"

out = "Intercompany_IMPORTRANGE_Queries.xlsx"
wb.save(out)
print(f"Saved: {out}")
